import openpyxl
import os

files = [
    'D:/Users/Public/php20250819/2026www/go-server/tests/comment_regression_safe_subset_results_20260331.xlsx',
    'D:/Users/Public/php20250819/2026www/go-server/tests/dx016_comment_moderation_test_20260327_112134_fixed.xlsx',
    'D:/Users/Public/php20250819/2026www/go-server/tests/two_site_multilang_rule_test_report_split_20260324_with_translation.xlsx',
]

for f in files:
    print('=== ' + os.path.basename(f) + ' ===')
    wb = openpyxl.load_workbook(f, read_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print('  Sheet:', sheet, ' 行数:', ws.max_row)
        for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
            print('  ', row)
    wb.close()
    print()

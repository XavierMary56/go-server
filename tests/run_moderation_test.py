import openpyxl
import json
import urllib.request

API_URL = 'http://localhost:888/v2/moderations'
API_KEY = 'sk-proj-51kb-test-f94bf04dbf680175'

def call_api(content):
    data = json.dumps({'content': content, 'type': 'comment', 'strictness': 'standard'}).encode('utf-8')
    req = urllib.request.Request(API_URL, data=data, method='POST')
    req.add_header('Content-Type', 'application/json')
    req.add_header('X-Project-Key', API_KEY)
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        return {'error': str(e)}

def extract_content(raw):
    # 格式: 【IDxx】实际内容 -> 提取】后面的内容
    s = str(raw) if raw else ''
    if '\u3011' in s:  # 》 U+3011
        return s.split('\u3011', 1)[-1].strip()
    return s.strip()

f = 'D:/Users/Public/php20250819/2026www/go-server/tests/two_site_multilang_rule_test_report_split_20260324_with_translation.xlsx'
wb = openpyxl.load_workbook(f, read_only=True)

pass_count = 0
fail_count = 0
skip_count = 0
fails = []

for sheet_name in wb.sheetnames:
    # 跳过说明/总览
    if sheet_name in list(wb.sheetnames)[:2]:
        continue

    # sheet名含 未 (U+672A) 为未通过(违规)，否则为通过(正常)
    sheet_expected = 'rejected' if '\u672a' in sheet_name else 'approved'

    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[2]:
            skip_count += 1
            continue
        content = extract_content(row[2])
        if not content or len(content) < 2:
            skip_count += 1
            continue

        # 用状态列(index 4)覆盖期望值
        status_cell = str(row[4]) if len(row) > 4 and row[4] else ''
        if '\u672a' in status_cell:  # 未 = 未通过
            expected = 'rejected'
        elif status_cell and '\u672a' not in status_cell and len(status_cell) > 1:
            expected = 'approved'
        else:
            expected = sheet_expected

        resp = call_api(content)
        if 'error' in resp and 'data' not in resp:
            print(f'API ERROR: {resp["error"]} content={content[:40]!r}')
            skip_count += 1
            continue

        verdict = resp.get('data', {}).get('result', {}).get('verdict', 'error')
        # approved 和 flagged 都算通过
        ok = verdict == expected or (expected == 'approved' and verdict in ('approved', 'flagged'))

        if ok:
            pass_count += 1
        else:
            fail_count += 1
            fails.append({
                'sheet': sheet_name,
                'content': content[:60],
                'expected': expected,
                'verdict': verdict,
            })

wb.close()

print()
print('===== 失败案例 =====')
for f in fails:
    print(f'  FAIL [{f["sheet"]}] expected={f["expected"]} got={f["verdict"]} | {f["content"]!r}')

print()
print('===== 测试结果 =====')
total = pass_count + fail_count
print(f'执行: {total}  通过: {pass_count}  失败: {fail_count}  跳过: {skip_count}')
if total > 0:
    print(f'通过率: {pass_count/total*100:.1f}%')

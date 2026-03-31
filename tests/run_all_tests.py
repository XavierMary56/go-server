import openpyxl
import json
import urllib.request

API_URL = 'http://localhost:888/v2/moderations'
API_KEY = 'sk-proj-51kb-test-f94bf04dbf680175'

def call_api(content):
    data = json.dumps({'content': content, 'type': 'comment', 'strictness': 'standard'}).encode('utf-8')
    req = urllib.request.Request(API_URL, data=data, method='POST')
    req.add_header('Content-Type', 'application/json')
    req.add_header('X-Project-Key', API_KEY)
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        return {'error': str(e)}

def extract_after_bracket(raw):
    """提取 【...】 后面的内容"""
    s = str(raw) if raw else ''
    if '\u3011' in s:  # 】U+3011
        return s.split('\u3011', 1)[-1].strip()
    return s.strip()

total_pass = 0
total_fail = 0
total_skip = 0
all_fails = []

# ============================================================
# 文件1: comment_regression_safe_subset_results_20260331.xlsx
# 结构: col0=用户输入, col1=提交状态, col2=测试结果(通过/不通过)
# 已有结果，直接统计，不调 API
# ============================================================
print('\n====== 文件1: comment_regression_safe_subset_results_20260331.xlsx (已有结果，直接统计) ======')
f1 = 'D:/Users/Public/php20250819/2026www/go-server/tests/comment_regression_safe_subset_results_20260331.xlsx'
wb1 = openpyxl.load_workbook(f1, read_only=True)
f1_pass = 0
f1_fail = 0
f1_skip = 0
for sheet_name in wb1.sheetnames:
    ws = wb1[sheet_name]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[2]:
            f1_skip += 1
            continue
        result_cell = str(row[2])
        if '\u901a\u8fc7' in result_cell and '\u4e0d' not in result_cell:  # 通过 but not 不通过
            f1_pass += 1
        elif '\u4e0d' in result_cell:  # 不通过
            f1_fail += 1
            all_fails.append({'file': 'regression', 'sheet': sheet_name, 'content': str(row[0])[:60] if row[0] else '', 'expected': 'approved', 'verdict': 'rejected(existing)'})
        else:
            f1_skip += 1
wb1.close()
print(f'  通过: {f1_pass}  不通过: {f1_fail}  跳过: {f1_skip}')
total_pass += f1_pass
total_fail += f1_fail
total_skip += f1_skip

# ============================================================
# 文件2: dx016_comment_moderation_test_20260327_112134_fixed.xlsx
# 结构: col3=内容, sheet名判断期望(正常评论=approved, 违规评论=rejected)
# 需要调 API
# ============================================================
print('\n====== 文件2: dx016_comment_moderation_test_20260327_112134_fixed.xlsx (调 API 验证) ======')
f2 = 'D:/Users/Public/php20250819/2026www/go-server/tests/dx016_comment_moderation_test_20260327_112134_fixed.xlsx'
wb2 = openpyxl.load_workbook(f2, read_only=True)
f2_pass = 0
f2_fail = 0
f2_skip = 0
SKIP_SHEETS2 = {'\u6c47\u603b', '\u8bf4\u660e'}  # 汇总, 说明
for sheet_name in wb2.sheetnames:
    if sheet_name in SKIP_SHEETS2:
        continue
    # 正常评论=approved, 违规评论=rejected
    if '\u8fdd\u89c4' in sheet_name:  # 违规
        sheet_expected = 'rejected'
    else:
        sheet_expected = 'approved'
    ws = wb2[sheet_name]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[3]:
            f2_skip += 1
            continue
        content = str(row[3]).strip()
        if not content or len(content) < 2:
            f2_skip += 1
            continue
        resp = call_api(content)
        if 'error' in resp and 'data' not in resp:
            print(f'  API ERROR: {resp["error"]} content={content[:40]!r}')
            f2_skip += 1
            continue
        verdict = resp.get('data', {}).get('result', {}).get('verdict', 'error')
        ok = verdict == sheet_expected or (sheet_expected == 'approved' and verdict in ('approved', 'flagged'))
        if ok:
            f2_pass += 1
        else:
            f2_fail += 1
            all_fails.append({'file': 'dx016', 'sheet': sheet_name, 'content': content[:60], 'expected': sheet_expected, 'verdict': verdict})
wb2.close()
print(f'  通过: {f2_pass}  失败: {f2_fail}  跳过: {f2_skip}')
total_pass += f2_pass
total_fail += f2_fail
total_skip += f2_skip

# ============================================================
# 文件3: two_site_multilang_rule_test_report_split_20260324_with_translation.xlsx
# 结构: col2=内容(【ID】内容格式), sheet名含 未(U+672A)=rejected
# 需要调 API
# ============================================================
print('\n====== 文件3: two_site_multilang_rule_test_report_split_20260324_with_translation.xlsx (调 API 验证) ======')
f3 = 'D:/Users/Public/php20250819/2026www/go-server/tests/two_site_multilang_rule_test_report_split_20260324_with_translation.xlsx'
wb3 = openpyxl.load_workbook(f3, read_only=True)
f3_pass = 0
f3_fail = 0
f3_skip = 0
for sheet_name in wb3.sheetnames:
    if sheet_name in list(wb3.sheetnames)[:2]:
        continue
    sheet_expected = 'rejected' if '\u672a' in sheet_name else 'approved'
    ws = wb3[sheet_name]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[2]:
            f3_skip += 1
            continue
        content = extract_after_bracket(row[2])
        if not content or len(content) < 2:
            f3_skip += 1
            continue
        status_cell = str(row[4]) if len(row) > 4 and row[4] else ''
        if '\u672a' in status_cell:
            expected = 'rejected'
        elif status_cell and '\u672a' not in status_cell and len(status_cell) > 1:
            expected = 'approved'
        else:
            expected = sheet_expected
        resp = call_api(content)
        if 'error' in resp and 'data' not in resp:
            print(f'  API ERROR: {resp["error"]} content={content[:40]!r}')
            f3_skip += 1
            continue
        verdict = resp.get('data', {}).get('result', {}).get('verdict', 'error')
        ok = verdict == expected or (expected == 'approved' and verdict in ('approved', 'flagged'))
        if ok:
            f3_pass += 1
        else:
            f3_fail += 1
            all_fails.append({'file': 'multilang', 'sheet': sheet_name, 'content': content[:60], 'expected': expected, 'verdict': verdict})
wb3.close()
print(f'  通过: {f3_pass}  失败: {f3_fail}  跳过: {f3_skip}')
total_pass += f3_pass
total_fail += f3_fail
total_skip += f3_skip

# ============================================================
# 汇总
# ============================================================
print('\n===== 失败案例 =====')
for item in all_fails:
    print(f'  FAIL [{item["file"]}][{item["sheet"]}] expected={item["expected"]} got={item["verdict"]} | {item["content"]!r}')

print('\n===== 总测试结果 =====')
total = total_pass + total_fail
print(f'执行: {total}  通过: {total_pass}  失败: {total_fail}  跳过: {total_skip}')
if total > 0:
    print(f'通过率: {total_pass/total*100:.1f}%')

import openpyxl
import json
import urllib.request
import urllib.parse

# 读取 two_site_multilang 文件中的测试内容（英文内容较多，编码问题小）
f = 'D:/Users/Public/php20250819/2026www/go-server/tests/two_site_multilang_rule_test_report_split_20260324_with_translation.xlsx'
wb = openpyxl.load_workbook(f, read_only=True)

test_cases = []
for sheet in wb.sheetnames:
    if sheet in ('说明', '总览'):
        continue
    ws = wb[sheet]
    expected = 'approved' if '通过' in sheet or '_通过' in sheet else 'rejected'
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 找content列（第3列或第4列）
        content = None
        for cell in row:
            if cell and isinstance(cell, str) and len(cell) > 3 and not cell.startswith('【'):
                content = cell
                break
        if content and len(content) > 2:
            test_cases.append({'content': content, 'expected': expected, 'sheet': sheet})

wb.close()

print(f'共收集到 {len(test_cases)} 条测试用例')
print('样本:')
for tc in test_cases[:5]:
    print(f'  [{tc["sheet"]}] expected={tc["expected"]} content={tc["content"][:60]}')

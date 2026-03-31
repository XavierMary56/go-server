import openpyxl

# 读取 dx016 违规测试 sheet，找英文内容
f = 'D:/Users/Public/php20250819/2026www/go-server/tests/dx016_comment_moderation_test_20260327_112134_fixed.xlsx'
wb = openpyxl.load_workbook(f, read_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'=== Sheet: {sheet_name} ===')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        print(f'  row{i+1}:', row)
    print()

wb.close()

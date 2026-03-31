import openpyxl

# two_site_multilang 文件里有英文内容，直接提取
f = 'D:/Users/Public/php20250819/2026www/go-server/tests/two_site_multilang_rule_test_report_split_20260324_with_translation.xlsx'
wb = openpyxl.load_workbook(f, read_only=True)

for sheet_name in wb.sheetnames:
    if sheet_name in wb.sheetnames[:2]:
        continue
    ws = wb[sheet_name]
    print(f'=== {sheet_name} (max_row={ws.max_row}) ===')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True)):
        print(f'  {row}')
    print()

wb.close()

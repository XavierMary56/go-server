import openpyxl
import json
import urllib.request

API_URL = 'http://localhost:888/v2/moderations'
API_KEY = 'sk-proj-51kb-test-f94bf04dbf680175'

def call_api(content):
    data = json.dumps({'content': content, 'type': 'comment', 'strictness': 'standard'}).encode('utf-8')
    req = urllib.request.Request(API_URL, data=data, method='POST')
    req.add_header('Content-Type', 'application/json')
    req.add_header('X-Project-Key', API_KEY)
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        return {'error': str(e)}

def extract_content(raw):
    if raw and '\u300b' in str(raw):
        return str(raw).split('\u300b', 1)[-1].strip()
    return str(raw).strip() if raw else ''

f = 'D:/Users/Public/php20250819/2026www/go-server/tests/two_site_multilang_rule_test_report_split_20260324_with_translation.xlsx'
wb = openpyxl.load_workbook(f, read_only=True)

# 先打印所有 sheet 的列头和前3行数据，确认列结构
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'=== {sheet_name} ===')
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        print(' ', row)
    print()

wb.close()
